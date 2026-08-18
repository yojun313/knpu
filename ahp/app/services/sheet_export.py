"""Excel 재현 패키지 — 계층·설문지·원자료·결과를 한 파일에 묶는다(PLAN.md 6.4).
몇 달 뒤에도 이 파일 하나로 무엇을 물었고 무엇을 받았는지, 어떻게 계산했는지
전부 재구성할 수 있어야 한다는 게 "재현 패키지"의 요구조건이다.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services.csv_schema import CSV_COLUMNS, format_value
from app.services.ahp_calc import pair_id


def _header(ws, columns):
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_response_rows(
    matrices: list[dict], nodes_by_uuid: dict, submissions_by_respondent: dict
) -> list[list]:
    """csv_schema.CSV_COLUMNS 순서(respondent,parent,item_a,item_b,value)의 long-format 행들.
    xlsx의 '응답 원자료' 시트와 CSV 내보내기가 이 함수 하나를 같이 쓴다 — 둘이
    포맷이 어긋나면 반입(entry_routes.import_csv)과 내보내기가 서로 안 맞게 된다.
    """
    rows = []
    for respondent_id, answers in submissions_by_respondent.items():
        label = respondent_id
        for m in matrices:
            child_uuids = m["child_uuids"]
            pairs = answers.get(m["matrix_id"], {})
            parent_name = nodes_by_uuid.get(m["parent_uuid"], {}).get("name", "")
            for i in range(len(child_uuids)):
                for j in range(i + 1, len(child_uuids)):
                    a, b = child_uuids[i], child_uuids[j]
                    pid = pair_id(a, b)
                    if pid not in pairs:
                        continue
                    stored = pairs[pid]
                    lo, _hi = sorted([a, b])
                    a_over_b = stored if a == lo else (1.0 / stored)
                    rows.append(
                        [
                            label,
                            parent_name,
                            nodes_by_uuid.get(a, {}).get("name", a),
                            nodes_by_uuid.get(b, {}).get("name", b),
                            format_value(a_over_b),
                        ]
                    )
    return rows


def build_workbook(
    project: dict,
    hierarchy: dict,
    survey: dict,
    nodes_by_uuid: dict,
    response_rows: list[list],
    results: dict | None,
) -> io.BytesIO:
    wb = Workbook()

    ws = wb.active
    ws.title = "프로젝트"
    _header(ws, ["항목", "값"])
    ws.append(["제목", project.get("title", "")])
    ws.append(["설명", project.get("description", "")])
    ws.append(["상태", project.get("status", "")])
    for k, v in (project.get("settings") or {}).items():
        ws.append([f"설정.{k}", v])

    ws2 = wb.create_sheet("계층")
    _header(ws2, ["uuid", "부모uuid", "이름", "설명", "순서", "깊이"])
    for n in hierarchy.get("nodes", []):
        ws2.append(
            [
                n["uuid"],
                n.get("parent_id") or "",
                n["name"],
                n.get("description", ""),
                n.get("order", 0),
                n.get("level", 0),
            ]
        )

    ws3 = wb.create_sheet("설문지")
    _header(ws3, ["matrix_id", "기준(부모)", "하위 항목들", "질문 문구"])
    for m in survey.get("matrices", []):
        children = ", ".join(
            nodes_by_uuid.get(c, {}).get("name", c) for c in m["child_uuids"]
        )
        parent_name = nodes_by_uuid.get(m["parent_uuid"], {}).get("name", "")
        ws3.append([m["matrix_id"], parent_name, children, m.get("question_text", "")])

    ws4 = wb.create_sheet("응답 원자료")
    _header(ws4, CSV_COLUMNS)
    for row in response_rows:
        ws4.append(row)

    if results:
        ws5 = wb.create_sheet("결과")
        _header(ws5, ["노드", "지역 가중치(소속 그룹 내)", "전역 가중치"])
        node_names = results.get("node_names", {})
        global_w = results.get("global_weights", {})
        local_w = results.get("local_weights", {})
        local_flat = {}
        for matrix_id, weights in local_w.items():
            for nid, w in weights.items():
                local_flat[nid] = w
        for nid, gw in sorted(global_w.items(), key=lambda kv: -kv[1]):
            ws5.append(
                [node_names.get(nid, nid), local_flat.get(nid, ""), round(gw, 4)]
            )

        ws6 = wb.create_sheet("개인별 CR")
        _header(ws6, ["응답자ID", "기준", "CR"])
        for rid, per_matrix in results.get("per_respondent_cr", {}).items():
            for matrix_id, cr in per_matrix.items():
                parent_uuid = next(
                    (
                        m["parent_uuid"]
                        for m in survey["matrices"]
                        if m["matrix_id"] == matrix_id
                    ),
                    matrix_id,
                )
                ws6.append(
                    [
                        rid,
                        node_names.get(parent_uuid, parent_uuid),
                        round(cr, 4) if cr is not None else "미완료",
                    ]
                )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
